# -*- coding: utf-8 -*-
"""
Excel Parser Module
负责多 Sheet Excel 读取与动态表头解析

设计原则:
1. 防御性编程：处理空值、非法字符及缺失 Sheet 的边界情况
2. 动态表头：基于 rules.yaml 中的 core_fields 别名匹配
3. 可读性：复杂聚合环节附带详细注释
"""

import re
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import yaml
from openpyxl import load_workbook

# 抑制 openpyxl 的 DataValidation 警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def _excel_serial_to_date(val: Any) -> str:
    """
    Excel 日期序列号转 YYYY-MM-DD。
    1900-01-01 为 1，46315 约为 2026-10-xx。
    """
    if val is None:
        return ""
    if isinstance(val, (int, float)) and val > 0 and val < 2958466:
        base = datetime(1899, 12, 30)  # Excel 基准
        d = base + timedelta(days=int(val))
        return d.strftime("%Y-%m-%d")
    return str(val).strip()


def _is_excel_serial_column(col: str) -> bool:
    """
    判断列名是否为 Excel 日期序列号（误作为列名）。
    当 Excel 表头行某单元格为日期时，pandas 可能读出为 46315 等数字字符串。
    """
    s = str(col).strip()
    if not s or not s.isdigit():
        return False
    try:
        n = int(s)
        return 1 <= n < 2958466
    except ValueError:
        return False


class ExcelParserError(Exception):
    """Excel 解析异常基类"""
    pass


class SheetNotFoundError(ExcelParserError):
    """Sheet 不存在异常"""
    pass


class RequiredFieldMissingError(ExcelParserError):
    """必填字段缺失异常"""
    pass


class ExcelParser:
    """
    Excel 解析器
    
    职责：
    - 读取多 Sheet Excel 文件
    - 动态表头解析（基于 core_fields 配置的别名匹配）
    - 数据清洗与空值处理
    - 生成符合 report_schema.md 的中间态数据
    """
    
    # 协议版本号
    PROTOCOL_VERSION = "2.1.0"
    
    # 非法字符正则（控制字符和不可见字符）
    ILLEGAL_CHAR_PATTERN = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
    
    def __init__(self, config_path: str = "config/rules.yaml"):
        """
        初始化解析器
        
        Args:
            config_path: 规则配置文件路径
        """
        self.config_path = Path(config_path)
        self._config: dict = {}
        self._core_fields: dict = {}
        self._priority_rules: dict = {}
        self._action_library: dict = {}
        self._high_risk_keywords: list = []
        self._sheet_column_mapping: dict = {}
        self._default_columns: list = []
        
        self._load_config()
    
    def _load_config(self) -> None:
        """加载并解析规则配置文件"""
        if not self.config_path.exists():
            raise FileNotFoundError(f"配置文件不存在: {self.config_path}")
        
        with open(self.config_path, 'r', encoding='utf-8') as f:
            self._config = yaml.safe_load(f)
        
        # 提取各配置模块
        self._core_fields = self._config.get('core_fields', {})
        self._priority_rules = self._config.get('priority_rules', {})
        self._action_library = self._config.get('action_library', {})
        self._high_risk_keywords = self._config.get('high_risk_keywords', [])
        self._sheet_column_mapping = self._config.get('sheet_column_mapping', {})
        self._default_columns = self._config.get('default_columns', [])
        self._implementation_summary_config = self._config.get('implementation_summary', {})
    
    def get_sheets(self) -> list[str]:
        """
        获取所有已解析的 Sheet 名称列表
        
        Returns:
            Sheet 名称列表（按优先级排序）
        """
        # 返回优先级规则中定义的 Sheet，按优先级排序
        sorted_sheets = sorted(
            self._priority_rules.keys(),
            key=lambda x: self._priority_rules.get(x, 999)
        )
        return sorted_sheets
    
    def parse(self, excel_path: str) -> dict:
        """
        解析 Excel 文件，生成符合 report_schema.md 的中间态数据
        
        Args:
            excel_path: Excel 文件路径
            
        Returns:
            符合 report_schema.md 规范的字典数据
        """
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")
        
        # 使用 openpyxl 获取所有 Sheet 名称
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        available_sheets = workbook.sheetnames
        workbook.close()
        
        # 【实施总表】解析第一个 Sheet 作为 implementation_summary
        # 策略：first_sheet=固定第一个 Sheet；name_match=按名称匹配
        implementation_summary = self._parse_implementation_summary(
            excel_file, available_sheets
        )
        
        # 确定实施总表使用的 Sheet 名称，该 Sheet 不进入 sections
        impl_summary_sheet_name = implementation_summary.get('sheet_name', '')
        
        # 解析结果容器
        sections = []
        all_external_links = []
        risk_alerts = []
        total_tasks = 0
        high_risk_count = 0
        
        # 按优先级顺序处理每个 Sheet，排除已作为实施总表的 Sheet
        sheets_to_process = [
            sheet for sheet in self._priority_rules.keys()
            if sheet in available_sheets and sheet != impl_summary_sheet_name
        ]
        sheets_to_process.sort(key=lambda x: self._priority_rules.get(x, 999))
        
        for sheet_name in sheets_to_process:
            section_data = self._parse_sheet(excel_file, sheet_name)
            
            if section_data:
                sections.append(section_data)
                total_tasks += section_data.get('task_count', 0)
                
                # 收集风险告警
                for action_group in section_data.get('action_groups', []):
                    if action_group.get('is_high_risk'):
                        risk_alerts.append({
                            'sheet_name': sheet_name,
                            'action_type': action_group.get('action_type'),
                            'task_count': action_group.get('task_count', 0),
                            'task_names': [
                                t.get('cells', [''])[0] for t in action_group.get('tasks', [])
                            ]
                        })
                        high_risk_count += action_group.get('task_count', 0)
        
        # 构建最终输出
        result = {
            'meta': {
                'source_file': excel_file.name,
                'generated_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
                'version': self.PROTOCOL_VERSION
            },
            'summary': {
                'total_tasks': total_tasks,
                'total_sheets': len(sections),
                'high_risk_count': high_risk_count,
                'has_external_links': len(all_external_links) > 0,
                'external_links': all_external_links
            },
            'has_risk_alerts': len(risk_alerts) > 0,
            'risk_alerts': risk_alerts,
            'implementation_summary': implementation_summary,
            'sections': sections
        }
        
        return result
    
    def _parse_implementation_summary(
        self, excel_file: Path, available_sheets: list[str]
    ) -> dict:
        """
        解析实施总表（Excel 第一个 Sheet 或按名称匹配的 Sheet）
        
        按 rules 中 implementation_summary 配置：
        - 过滤 Unnamed、空列名
        - 列映射到 output_columns
        - 日期列做 Excel 序列号 → YYYY-MM-DD 转换
        - 无序号列时自动生成 1,2,3...
        
        Returns:
            implementation_summary 字典，columns 固定为 output_columns 顺序
        """
        empty_result = {
            'sheet_name': '',
            'columns': [],
            'rows': [],
            'has_data': False
        }
        
        if not available_sheets:
            return empty_result
        
        strategy = self._implementation_summary_config.get('strategy', 'first_sheet')
        sheet_names = self._implementation_summary_config.get('sheet_names', [])
        output_columns = self._implementation_summary_config.get(
            'output_columns', ['序号', '任务', '开始时间', '结束时间', '实施人', '复核人']
        )
        column_mapping = self._implementation_summary_config.get('column_mapping', {})
        date_columns = set(self._implementation_summary_config.get('date_columns', ['开始时间', '结束时间']))
        auto_sequence = self._implementation_summary_config.get('auto_sequence', True)
        drop_unnamed = self._implementation_summary_config.get('drop_unnamed_columns', True)
        
        target_sheet = None
        if strategy == 'first_sheet':
            target_sheet = available_sheets[0]
        elif strategy == 'name_match' and sheet_names:
            for name in sheet_names:
                if name in available_sheets:
                    target_sheet = name
                    break
        
        if not target_sheet:
            return empty_result
        
        try:
            df = pd.read_excel(excel_file, sheet_name=target_sheet, header=0)
            
            if df.empty:
                return {
                    'sheet_name': target_sheet,
                    'columns': list(output_columns),
                    'rows': [],
                    'has_data': False
                }
            
            df = self._clean_dataframe(df)
            
            if df.empty:
                return {
                    'sheet_name': target_sheet,
                    'columns': list(output_columns),
                    'rows': [],
                    'has_data': False
                }
            
            # 1. 过滤 Unnamed、空列名、Excel 日期序列号（46315 等误作列名）
            cols_raw = [str(c).strip() if pd.notna(c) else '' for c in df.columns.tolist()]
            if drop_unnamed:
                keep_idx = [
                    i for i, c in enumerate(cols_raw)
                    if c
                    and not (c.lower().startswith('unnamed') or c == '')
                    and not _is_excel_serial_column(c)
                ]
                df = df.iloc[:, keep_idx]
                cols_raw = [str(c).strip() if pd.notna(c) else '' for c in df.columns.tolist()]
            
            excel_cols = cols_raw
            
            # 2. 建立标准列 -> Excel 列 映射
            # column_mapping: { "序号": ["任务序号","序号",...], "任务": ["任务名",...], ... }
            std_to_excel: dict[str, Optional[str]] = {}
            for std_col in output_columns:
                aliases = column_mapping.get(std_col, [std_col])
                if isinstance(aliases, str):
                    aliases = [aliases]
                found = None
                for ec in excel_cols:
                    ec_norm = ec.strip()
                    for alias in aliases:
                        if str(alias).strip() == ec_norm:
                            found = ec
                            break
                    if found:
                        break
                std_to_excel[std_col] = found
            
            # 3. 构建行数据：按 output_columns 顺序
            rows = []
            for row_idx, row in df.iterrows():
                cells = []
                for std_col in output_columns:
                    excel_col = std_to_excel.get(std_col)
                    if std_col == '序号':
                        if excel_col and excel_col in row.index:
                            val = row.get(excel_col, '')
                            cells.append('' if pd.isna(val) else self._sanitize_string(str(val)))
                        elif auto_sequence:
                            cells.append(str(row_idx + 1))
                        else:
                            cells.append('')
                    elif excel_col and excel_col in row.index:
                        val = row.get(excel_col, '')
                        if pd.isna(val):
                            cells.append('')
                        elif std_col in date_columns:
                            cells.append(_excel_serial_to_date(val))
                        else:
                            cells.append(self._sanitize_string(str(val)))
                    else:
                        cells.append('')
                rows.append({'cells': cells})
            
            return {
                'sheet_name': target_sheet,
                'columns': list(output_columns),
                'rows': rows,
                'has_data': len(rows) > 0
            }
        except Exception as e:
            print(f"警告: 解析实施总表 '{target_sheet}' 时出错: {e}")
            return {
                'sheet_name': target_sheet,
                'columns': list(output_columns),
                'rows': [],
                'has_data': False
            }
    
    def _parse_sheet(self, excel_file: Path, sheet_name: str) -> Optional[dict]:
        """
        解析单个 Sheet
        
        Args:
            excel_file: Excel 文件路径对象
            sheet_name: Sheet 名称
            
        Returns:
            章节数据字典，或 None（如果 Sheet 为空）
        """
        try:
            # 读取 Sheet 数据
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0)
            
            # 空数据处理
            if df.empty:
                return None
            
            # 清洗数据：处理空值和非法字符
            df = self._clean_dataframe(df)
            
            # 动态表头解析：建立列名到核心字段的映射
            field_mapping = self._build_field_mapping(df.columns.tolist())
            
            # 验证必填字段
            self._validate_required_fields(field_mapping)
            
            # 提取核心字段数据
            action_column = field_mapping.get('action_type')
            task_name_column = field_mapping.get('task_name')
            deploy_unit_column = field_mapping.get('deploy_unit')
            executor_column = field_mapping.get('executor')
            external_link_column = field_mapping.get('external_link')
            
            # 按 action_type 分组聚合
            # 聚合算法：相同 Sheet 下相同操作类型的任务合并为一组
            action_groups = {}
            
            for idx, row in df.iterrows():
                # 获取操作类型
                action_type = self._safe_get_value(row, action_column)
                if not action_type:
                    continue  # 跳过无操作类型的行
                
                action_type = str(action_type).strip()
                
                # 初始化操作组
                if action_type not in action_groups:
                    action_groups[action_type] = []
                
                # 构建任务数据
                task_data = {
                    'task_name': self._safe_get_value(row, task_name_column, ''),
                    'deploy_unit': self._safe_get_value(row, deploy_unit_column, ''),
                    'executor': self._safe_get_value(row, executor_column, ''),
                    'external_link': self._safe_get_value(row, external_link_column, ''),
                    'raw_data': self._extract_raw_data(row)
                }
                
                action_groups[action_type].append(task_data)
            
            # 构建 action_groups 列表
            # 获取该章节的列定义（用于 cells 提取）
            columns = self._get_columns_for_sheet(sheet_name)
            
            formatted_action_groups = []
            for action_type, tasks in action_groups.items():
                # 从 action_library 获取操作说明
                action_config = self._action_library.get(action_type, {})
                
                # 判断是否高危操作
                is_high_risk = self._is_high_risk(action_type)
                
                # 将任务数据转换为 cells 数组格式
                formatted_tasks = []
                for task in tasks:
                    # 按列顺序提取 cells 数组
                    cells = self._extract_cells_by_columns(
                        task.get('raw_data', {}), 
                        columns
                    )
                    formatted_tasks.append({'cells': cells})
                
                formatted_action_groups.append({
                    'action_type': action_type,
                    'instruction': action_config.get(
                        'instruction', 
                        f"执行以下{action_type}操作："
                    ),
                    'is_high_risk': is_high_risk,
                    'task_count': len(formatted_tasks),
                    'tasks': formatted_tasks
                })
            
            # 获取章节优先级
            priority = self._priority_rules.get(sheet_name, 999)
            
            return {
                'section_name': sheet_name,
                'priority': priority,
                'has_action_groups': len(formatted_action_groups) > 0,
                'columns': columns,
                'task_count': sum(len(g['tasks']) for g in formatted_action_groups),
                'action_groups': formatted_action_groups
            }
            
        except Exception as e:
            # 防御性编程：记录错误但不中断整个解析过程
            print(f"警告: 解析 Sheet '{sheet_name}' 时出错: {e}")
            return None
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        清洗 DataFrame：处理空值和非法字符
        
        Args:
            df: 原始 DataFrame
            
        Returns:
            清洗后的 DataFrame
        """
        # 清理列名中的空白字符
        df.columns = df.columns.map(lambda x: str(x).strip() if pd.notna(x) else '')
        
        # 移除完全空白的行
        df = df.dropna(how='all')
        
        # 移除所有列都为空字符串的行
        df = df.loc[~(df == '').all(axis=1)]
        
        # 重置索引
        df = df.reset_index(drop=True)
        
        return df
    
    def _build_field_mapping(self, columns: list[str]) -> dict[str, Optional[str]]:
        """
        建立核心字段到实际列名的映射（动态表头解析）
        
        通过 core_fields 中定义的别名，在 Excel 列名中查找匹配项
        支持大小写不敏感和部分匹配
        
        Args:
            columns: Excel 列名列表
            
        Returns:
            字段映射字典 {field_key: actual_column_name}
        """
        mapping = {}
        
        for field_key, field_config in self._core_fields.items():
            aliases = field_config.get('aliases', [])
            actual_column = None
            
            # 精确匹配（忽略大小写和空白）
            for col in columns:
                col_normalized = str(col).strip().lower()
                for alias in aliases:
                    if col_normalized == alias.lower().strip():
                        actual_column = col
                        break
                if actual_column:
                    break
            
            # 如果精确匹配失败，尝试包含匹配
            if not actual_column:
                for col in columns:
                    col_normalized = str(col).strip().lower()
                    for alias in aliases:
                        if alias.lower().strip() in col_normalized:
                            actual_column = col
                            break
                    if actual_column:
                        break
            
            mapping[field_key] = actual_column
        
        return mapping
    
    def _validate_required_fields(self, field_mapping: dict[str, Optional[str]]) -> None:
        """
        验证必填字段是否存在
        
        Args:
            field_mapping: 字段映射字典
            
        Raises:
            RequiredFieldMissingError: 必填字段缺失时抛出
        """
        missing_fields = []
        
        for field_key, field_config in self._core_fields.items():
            if field_config.get('required', False):
                if not field_mapping.get(field_key):
                    missing_fields.append(field_key)
        
        if missing_fields:
            raise RequiredFieldMissingError(
                f"必填字段缺失: {', '.join(missing_fields)}"
            )
    
    def _safe_get_value(
        self, 
        row: pd.Series, 
        column_name: Optional[str], 
        default: str = ''
    ) -> str:
        """
        安全获取单元格值，处理空值和非法字符
        
        Args:
            row: 数据行
            column_name: 列名
            default: 默认值
            
        Returns:
            清洗后的字符串值
        """
        if not column_name:
            return default
        
        value = row.get(column_name, default)
        
        # 处理 NaN
        if pd.isna(value):
            return default
        
        # 转换为字符串并清理非法字符
        str_value = str(value).strip()
        str_value = self._sanitize_string(str_value)
        
        return str_value if str_value else default
    
    def _sanitize_string(self, text: str) -> str:
        """
        清理字符串中的非法字符（控制字符等）
        
        Args:
            text: 原始字符串
            
        Returns:
            清理后的字符串
        """
        # 移除控制字符
        cleaned = self.ILLEGAL_CHAR_PATTERN.sub('', text)
        
        # 将连续空白替换为单个空格
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        return cleaned.strip()
    
    def _extract_raw_data(self, row: pd.Series) -> dict:
        """
        提取原始行数据的完整字典（保留所有列）
        
        Args:
            row: 数据行
            
        Returns:
            原始数据字典
        """
        raw_data = {}
        for col in row.index:
            value = row[col]
            if pd.isna(value):
                raw_data[col] = ''
            else:
                raw_data[col] = self._sanitize_string(str(value))
        return raw_data
    
    def _is_high_risk(self, action_type: str) -> bool:
        """
        判断操作是否为高危操作
        
        Args:
            action_type: 操作类型
            
        Returns:
            是否高危
        """
        # 检查是否在 action_library 中标记为高危
        action_config = self._action_library.get(action_type, {})
        if action_config.get('is_high_risk', False):
            return True
        
        # 检查是否包含高危关键字
        for keyword in self._high_risk_keywords:
            if keyword in action_type:
                return True
        
        return False
    
    def _extract_cells_by_columns(self, raw_data: dict, columns: list[str]) -> list[str]:
        """
        从原始数据中按配置的列顺序提取单元格值
        
        Args:
            raw_data: 原始行数据字典 {列名: 值}
            columns: 列名列表（定义顺序）
            
        Returns:
            按列顺序排列的单元格值列表
        """
        cells = []
        for col in columns:
            # 从 raw_data 中按列名获取值，默认为空字符串
            value = raw_data.get(col, '')
            cells.append(str(value) if value else '')
        return cells
    
    def _get_columns_for_sheet(self, sheet_name: str) -> list[str]:
        """
        获取指定 Sheet 应展示的列名
        
        Args:
            sheet_name: Sheet 名称
            
        Returns:
            列名列表
        """
        if sheet_name in self._sheet_column_mapping:
            return self._sheet_column_mapping[sheet_name].get('columns', [])
        return self._default_columns
    
    def get_columns_for_sheet(self, sheet_name: str) -> list[str]:
        """
        获取指定 Sheet 应展示的列名（公开接口）
        
        Args:
            sheet_name: Sheet 名称
            
        Returns:
            列名列表
        """
        return self._get_columns_for_sheet(sheet_name)


# 便捷函数
def parse_excel(excel_path: str, config_path: str = "config/rules.yaml") -> dict:
    """
    解析 Excel 文件的便捷函数
    
    Args:
        excel_path: Excel 文件路径
        config_path: 规则配置文件路径
        
    Returns:
        符合 report_schema.md 规范的字典数据
    """
    parser = ExcelParser(config_path)
    return parser.parse(excel_path)
